import openpyxl
from flask import Flask, request
from twilio.twiml.messaging_response import MessagingResponse

app = Flask(__name__)

def leer_excel_local(nombre_pestana):
    try:
        wb = openpyxl.load_workbook("menu_y_promos_comercio.xlsx", data_only=True)
        pestana_encontrada = None
        for name in wb.sheetnames:
            if name.strip().lower() == nombre_pestana.strip().lower():
                pestana_encontrada = name
                break
        
        if pestana_encontrada:
            sheet = wb[pestana_encontrada]
            filas = []
            for row in sheet.iter_rows(values_only=True):
                fila_str = [str(cell).strip() if cell is not None else "" for cell in row]
                if any(fila_str):
                    filas.append(fila_str)
            return filas[1:] if len(filas) > 1 else []
    except Exception as e:
        print(f"Error al leer el excel local: {e}")
    return []

sesiones_usuarios = {}

@app.route("/")
def home():
    return "¡El Bot de Pedidos de la Pizzería está activo y funcionando correctamente! 🍕🤖"

@app.route("/webhook", methods=['POST'])
def whatsapp_webhook():
    incoming_msg = request.values.get('Body', '').strip().lower()
    sender_number = request.values.get('From', '')
    
    resp = MessagingResponse()
    msg = resp.message()

    if sender_number not in sesiones_usuarios:
        sesiones_usuarios[sender_number] = {"estado": "INICIO", "carrito": []}
    
    usuario = sesiones_usuarios[sender_number]
    estado_actual = usuario["estado"]

    if "hola" in incoming_msg or "menu" in incoming_msg or "inicio" in incoming_msg:
        usuario["estado"] = "MENU"
        usuario["carrito"] = []
        msg.body(
            "¡Hola! 👋 Bienvenido a nuestro servicio de pedidos automáticos.\n\n"
            "¿Qué te gustaría hacer hoy?\n"
            "1️⃣ Ver Catálogo de Productos\n"
            "2️⃣ Ver Promociones y Combos\n"
            "3️⃣ Ver mi Carrito actual\n\n"
            "Respondé con el número de la opción."
        )

    elif estado_actual == "MENU":
        if incoming_msg == "1":
            filas = leer_excel_local("Menú y Productos")
            if filas:
                texto = "📋 *CATÁLOGO DE PRODUCTOS*:\n\n"
                for r in filas[:10]:
                    if len(r) >= 5:
                        codigo = r[0]   # Columna A: Código
                        producto = r[2] # Columna C: Producto / Variedad
                        precio = r[4]   # Columna E: Precio ($)
                        if codigo:
                            texto += f"🔹 *[{codigo}]* {producto} - ${precio}\n"
                texto += "\nRespondé con el *Código* del producto (ej: P01) para sumarlo, o *0* para volver."
                usuario["estado"] = "ESPERANDO_PRODUCTO"
                msg.body(texto)
            else:
                msg.body("⚠️ No se pudieron leer los productos del archivo Excel local.")

        elif incoming_msg == "2":
            filas = leer_excel_local("Promociones y Combos")
            if filas:
                texto = "🔥 *PROMOCIONES Y COMBOS*:\n\n"
                for r in filas:
                    if len(r) >= 4:
                        codigo = r[0]
                        promo = r[1]
                        precio = r[3]
                        if codigo:
                            texto += f"⭐ *[{codigo}]* {promo} - *${precio}*\n"
                texto += "\nRespondé con el *Código* de la promo (ej: C01), o *0* para volver."
                usuario["estado"] = "ESPERANDO_PRODUCTO"
                msg.body(texto)
            else:
                msg.body("⚠️ No se pudieron leer las promociones.")

        elif incoming_msg == "3":
            if not usuario["carrito"]:
                msg.body("🛒 Tu carrito está vacío. Escribí *menu* para ver las opciones.")
            else:
                detalle = "🛒 *Tu Carrito Actual*:\n"
                total = 0
                for item in usuario["carrito"]:
                    detalle += f"- {item['nombre']}: ${item['precio']}\n"
                    limpio = item['precio'].replace('.', '').replace(',', '.')
                    total += float(limpio) if limpio.replace('.', '', 1).isdigit() else 0
                detalle += f"\nTotal a pagar: *${total}*\n\nRespondé *CONFIRMAR* para finalizar o *MENU* para seguir comprando."
                usuario["estado"] = "CONFIRMACION"
                msg.body(detalle)
        else:
            msg.body("Opción no válida. Por favor respondé 1, 2 o 3.")

    elif estado_actual == "ESPERANDO_PRODUCTO":
        if incoming_msg == "0":
            usuario["estado"] = "MENU"
            msg.body("Volviste al menú principal. Escribí 1, 2 o 3.")
        else:
            encontrado = None
            for r in leer_excel_local("Menú y Productos"):
                if len(r) >= 5 and r[0].lower() == incoming_msg:
                    encontrado = {"nombre": r[2], "precio": r[4]}
                    break
            
            if not encontrado:
                for r in leer_excel_local("Promociones y Combos"):
                    if len(r) >= 4 and r[0].lower() == incoming_msg:
                        encontrado = {"nombre": r[1], "precio": r[3]}
                        break

            if encontrado:
                usuario["carrito"].append(encontrado)
                msg.body(f"✅ ¡Agregado: *{encontrado['nombre']}* (${encontrado['precio']})!\n\n¿Querés otro producto (escribí su código) o ver tu carrito escribiendo *3*?")
            else:
                msg.body("❌ Código no encontrado. Verificá el código en el catálogo o escribí *0* para volver.")

    elif estado_actual == "CONFIRMACION":
        if "confirmar" in incoming_msg:
            msg.body("🎉 ¡Pedido confirmado con éxito! El local ya lo está preparando. ¡Muchas gracias! 🙌")
            usuario["carrito"] = []
            usuario["estado"] = "MENU"
        else:
            usuario["estado"] = "MENU"
            msg.body("Operación cancelada. Escribí *hola* para empezar de nuevo.")

    return str(resp)

if __name__ == "__main__":
    app.run(port=5000, debug=True) 
